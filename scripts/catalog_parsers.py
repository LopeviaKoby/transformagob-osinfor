"""
Parsers de fuentes raw para el catalogo SQLite.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any, Iterable
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader

from scripts.catalog_normalization import (
    canonicalize_species,
    clean_text,
    composite_balance_key,
    composite_log_key,
    composite_tree_key,
    normalize_date,
    normalize_decimal,
    normalize_gtf,
    normalize_header,
    normalize_log_code,
    normalize_parcela,
    normalize_species,
    normalize_title,
    normalize_tree_code,
    parent_tree_from_log_code,
    species_signature,
)


EXPECTED_SOURCE_PATHS = [
    "raw/censo/censo_forestal.xlsx",
    "raw/muestra/muestra_supervisada.xlsx",
    "raw/libro/pc01/libro_operaciones_pc01.xlsx",
    "raw/libro/pc01/libro_operaciones_pc02.xlsx",
    "raw/libro/pc01/libro_operaciones_pc03.xlsx",
    "raw/balance/pc01/balance_extraccion_pc01.pdf",
    "raw/balance/pc01/balance_extraccion_pc02.pdf",
    "raw/balance/pc01/balance_extraccion_pc03.pdf",
]

_WORKBOOK_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def resolve_repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def resolve_real_sources(repo_root: Path | None = None) -> list[Path]:
    repo_root = repo_root or resolve_repo_root()
    resolved: list[Path] = []
    raw_root = repo_root / "raw"

    for relative in EXPECTED_SOURCE_PATHS:
        candidate = repo_root / relative
        if candidate.exists():
            resolved.append(candidate)
            continue

        basename = Path(relative).name
        matches = sorted(raw_root.rglob(basename))
        if not matches:
            raise FileNotFoundError(
                f"No se encontro la fuente esperada por nombre: {relative}"
            )
        resolved.append(matches[0])

    return resolved


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def excel_sheet_names(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("xl/workbook.xml"))
    sheets = root.find("main:sheets", _WORKBOOK_NS)
    if sheets is None:
        return []
    return [sheet.attrib.get("name", "") for sheet in sheets]


def pdf_page_count(path: Path) -> int:
    return len(PdfReader(str(path)).pages)


def collect_source_inventory(repo_root: Path | None = None) -> dict[str, Any]:
    repo_root = repo_root or resolve_repo_root()
    sources: list[dict[str, Any]] = []
    for path in resolve_real_sources(repo_root):
        rel = path.relative_to(repo_root).as_posix()
        item = {
            "relative_path": rel,
            "size_bytes": path.stat().st_size,
            "sha256": file_sha256(path),
        }
        if path.suffix.lower() == ".xlsx":
            item["sheets"] = excel_sheet_names(path)
            item["pages"] = None
        else:
            item["sheets"] = None
            item["pages"] = pdf_page_count(path)
        sources.append(item)
    return {"repo_root": str(repo_root), "sources": sources}


def detect_header_row(rows: list[list[Any]], expected_terms: set[str]) -> int:
    best_index = -1
    best_score = -1
    for index, row in enumerate(rows[:20]):
        normalized_cells = [normalize_header(value) for value in row if clean_text(value)]
        if not normalized_cells:
            continue
        row_text = " ".join(normalized_cells)
        score = sum(1 for term in expected_terms if term in row_text)
        if score > best_score:
            best_score = score
            best_index = index

    if best_index == -1 or best_score < max(2, len(expected_terms) // 2):
        raise ValueError(
            f"No se pudo detectar encabezado con terminos {sorted(expected_terms)}"
        )

    return best_index


def build_header(row: list[Any]) -> list[str]:
    header: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row):
        normalized = normalize_header(value) or f"col_{index}"
        seen[normalized] = seen.get(normalized, 0) + 1
        if seen[normalized] > 1:
            normalized = f"{normalized}_{seen[normalized]}"
        header.append(normalized)
    return header


def row_to_mapping(header: list[str], row: list[Any]) -> dict[str, Any]:
    padded = row + [None] * max(0, len(header) - len(row))
    return {header[index]: padded[index] for index in range(len(header))}


def iter_sheet_rows(path: Path, sheet_name: str) -> tuple[list[list[Any]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return rows, len(rows)


def extract_sheet_metadata(rows: list[list[Any]]) -> dict[str, str | None]:
    metadata: dict[str, str | None] = {}
    for row in rows[:4]:
        if not row:
            continue
        key = normalize_header(row[0])
        non_empty = [clean_text(value) for value in row if clean_text(value)]
        if not key or len(non_empty) < 2:
            continue
        metadata[key] = non_empty[-1]
    return metadata


def _source_payload(relative_path: str, sheet: str | None, row_number: int | None, raw: dict[str, Any], normalized: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_relative_path": relative_path,
        "source_sheet": sheet,
        "source_row_number": row_number,
        "raw_payload_json": json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str),
        "normalized_payload_json": json.dumps(
            normalized,
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        ),
    }


def build_species_aliases(
    census_rows: Iterable[dict[str, Any]],
    supervision_rows: Iterable[dict[str, Any]],
) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for row in list(census_rows) + list(supervision_rows):
        for field in ("especie_norm", "especie_supervision_norm"):
            species = row.get(field)
            if not species:
                continue
            signature = species_signature(species)
            if signature:
                aliases[signature] = species
            left = species.split("|", 1)[0].strip() if "|" in species else species
            lead_tokens = re.split(r"\s+", left)
            if len(lead_tokens) >= 2:
                aliases[" ".join(lead_tokens[:2]).lower()] = species
    return aliases


def parse_census_workbook(path: Path, repo_root: Path | None = None) -> dict[str, Any]:
    repo_root = repo_root or resolve_repo_root()
    rows, _ = iter_sheet_rows(path, "CENSO")
    header_index = detect_header_row(rows, {"num_thabilitante", "codigo", "volumen", "pca"})
    header = build_header(rows[header_index])
    relative_path = path.relative_to(repo_root).as_posix()

    loaded: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        mapping = row_to_mapping(header, row)
        if not any(clean_text(value) for value in mapping.values()):
            continue

        title = normalize_title(mapping.get("num_thabilitante"))
        parcela = normalize_parcela(mapping.get("pca"))
        codigo = normalize_tree_code(mapping.get("codigo"))
        if not (title and parcela and codigo):
            rejected.append(
                {
                    "reason": "missing_identity",
                    "source_row_number": row_number,
                    "row": mapping,
                }
            )
            continue

        normalized = {
            "titulo_habilitante_norm": title,
            "titular_norm": clean_text(mapping.get("titular_actual")),
            "parcela_corta_norm": parcela,
            "codigo_arbol_norm": codigo,
            "composite_tree_key": composite_tree_key(title, parcela, codigo),
            "especie_norm": normalize_species(mapping.get("especies")),
            "volumen_censo_text": normalize_decimal(mapping.get("volumen")),
            "dap_text": normalize_decimal(mapping.get("dap")) if clean_text(mapping.get("dap")) else None,
            "ac_text": normalize_decimal(mapping.get("ac")) if clean_text(mapping.get("ac")) else None,
            "dmc_text": normalize_decimal(mapping.get("dmc")) if clean_text(mapping.get("dmc")) else None,
            "plan_operativo_norm": clean_text(mapping.get("nombre_poa")),
            "resolucion_norm": clean_text(mapping.get("aresolucion_num")),
            "fecha_resolucion": normalize_date(mapping.get("aresolucion_fecha")),
        }

        record = {
            "titulo_habilitante_original": clean_text(mapping.get("num_thabilitante")),
            "titular_original": clean_text(mapping.get("titular_actual")),
            "parcela_corta_original": clean_text(mapping.get("pca")),
            "codigo_arbol_original": clean_text(mapping.get("codigo")),
            "especie_original": clean_text(mapping.get("especies")),
            "plan_operativo_original": clean_text(mapping.get("nombre_poa")),
            "resolucion_original": clean_text(mapping.get("aresolucion_num")),
            **normalized,
            **_source_payload(relative_path, "CENSO", row_number, mapping, normalized),
        }
        loaded.append(record)

    return {"loaded": loaded, "rejected": rejected}


def parse_supervision_workbook(
    path: Path,
    canonical_species: dict[str, str] | None = None,
    repo_root: Path | None = None,
) -> dict[str, Any]:
    repo_root = repo_root or resolve_repo_root()
    rows, _ = iter_sheet_rows(path, "Hoja1")
    header_index = detect_header_row(
        rows,
        {"num_thabilitante", "codigo", "codigo_campo", "desc_especies", "pca_campo"},
    )
    header = build_header(rows[header_index])
    relative_path = path.relative_to(repo_root).as_posix()

    loaded: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        mapping = row_to_mapping(header, row)
        if not any(clean_text(value) for value in mapping.values()):
            continue

        title = normalize_title(mapping.get("num_thabilitante"))
        parcela = normalize_parcela(mapping.get("pca_campo") or mapping.get("pca"))
        codigo = normalize_tree_code(mapping.get("codigo_campo") or mapping.get("codigo"))
        if not (title and parcela and codigo):
            rejected.append(
                {
                    "reason": "missing_identity",
                    "source_row_number": row_number,
                    "row": mapping,
                }
            )
            continue

        raw_species = normalize_species(mapping.get("desc_especies_campo"))
        normalized_species = (
            canonicalize_species(raw_species, canonical_species or {})
            if raw_species
            else None
        )

        normalized = {
            "titulo_habilitante_norm": title,
            "parcela_corta_norm": parcela,
            "codigo_arbol_norm": codigo,
            "composite_tree_key": composite_tree_key(title, parcela, codigo),
            "especie_supervision_norm": normalized_species,
            "especie_censo_referida_norm": normalize_species(mapping.get("desc_especies")),
            "coincide_especies_norm": clean_text(mapping.get("coincide_especies")),
            "fecha_informe": normalize_date(mapping.get("fecha_informe")),
        }

        record = {
            "titulo_habilitante_original": clean_text(mapping.get("num_thabilitante")),
            "parcela_corta_original": clean_text(mapping.get("pca_campo") or mapping.get("pca")),
            "codigo_arbol_original": clean_text(mapping.get("codigo_campo") or mapping.get("codigo")),
            "especie_supervision_original": raw_species,
            "especie_censo_referida_original": normalize_species(mapping.get("desc_especies")),
            **normalized,
            **_source_payload(relative_path, "Hoja1", row_number, mapping, normalized),
        }
        loaded.append(record)

    return {"loaded": loaded, "rejected": rejected}


def _sheet_name_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return {sheet.title.lower(): sheet.title for sheet in workbook.worksheets}


def parse_operations_workbook(
    path: Path,
    canonical_species: dict[str, str] | None = None,
    repo_root: Path | None = None,
) -> dict[str, Any]:
    repo_root = repo_root or resolve_repo_root()
    relative_path = path.relative_to(repo_root).as_posix()
    sheet_map = _sheet_name_map(path)
    required = {"tala", "trozado", "despacho"}
    if not required.issubset(sheet_map):
        missing = sorted(required - set(sheet_map))
        raise ValueError(f"{relative_path}: faltan hojas obligatorias {missing}")

    result = {
        "fellings": [],
        "logs": [],
        "dispatches": [],
        "rejected": [],
        "metadata": {},
    }

    for logical_name in ("tala", "trozado", "despacho"):
        sheet_name = sheet_map[logical_name]
        rows, _ = iter_sheet_rows(path, sheet_name)
        metadata = extract_sheet_metadata(rows)
        result["metadata"][logical_name] = metadata
        header_index = detect_header_row(
            rows,
            (
                {"codigo_del_arbol", "fecha", "especie", "volumen"}
                if logical_name == "tala"
                else {"codigo_de_la_troza", "fecha", "especie", "volumen"}
                if logical_name == "trozado"
                else {"codigo_de_troza", "numero_de_gtf", "codigo_de_despacho"}
            ),
        )
        header = build_header(rows[header_index])

        title = normalize_title(metadata.get("titulo_habilitante"))
        resolution = clean_text(metadata.get("n_resolucion"))
        parcela = normalize_parcela(resolution)
        if not (title and parcela):
            raise ValueError(f"{relative_path}:{sheet_name} sin titulo/parcela interpretables")

        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            mapping = row_to_mapping(header, row)
            if not any(clean_text(value) for value in mapping.values()):
                continue

            try:
                if logical_name == "tala":
                    code = normalize_tree_code(mapping.get("codigo_del_arbol"))
                    if not code:
                        raise ValueError("missing_tree_code")
                    raw_species = clean_text(mapping.get("especie"))
                    normalized_species = canonicalize_species(
                        raw_species,
                        canonical_species or {},
                    )
                    normalized = {
                        "titulo_habilitante_norm": title,
                        "parcela_corta_norm": parcela,
                        "codigo_arbol_norm": code,
                        "composite_tree_key": composite_tree_key(title, parcela, code),
                        "especie_norm": normalized_species,
                        "fecha_operacion": normalize_date(mapping.get("fecha")),
                        "volumen_text": normalize_decimal(mapping.get("volumen_m3") or mapping.get("volumen_m3_2") or mapping.get("volumen_m3_3") or mapping.get("volumen_m3_4") or mapping.get("volumen_m3_5") or mapping.get("volumen_m3_6") or mapping.get("volumen_m3_7") or mapping.get("volumen_m3_8") or mapping.get("volumen_m3_9") or mapping.get("volumen")),
                        "r_private_value": clean_text(mapping.get("r")),
                    }
                    result["fellings"].append(
                        {
                            "titulo_habilitante_original": title,
                            "parcela_corta_original": parcela,
                            "codigo_arbol_original": clean_text(mapping.get("codigo_del_arbol")),
                            "especie_original": raw_species,
                            "diametro_mayor_text": normalize_decimal(mapping.get("diametro_mayor_m")) if clean_text(mapping.get("diametro_mayor_m")) else None,
                            "diametro_menor_text": normalize_decimal(mapping.get("diametro_menor_m")) if clean_text(mapping.get("diametro_menor_m")) else None,
                            "longitud_text": normalize_decimal(mapping.get("longitud_aprovechable_m")) if clean_text(mapping.get("longitud_aprovechable_m")) else None,
                            "observaciones_private": clean_text(mapping.get("observaciones")),
                            **normalized,
                            **_source_payload(relative_path, sheet_name, row_number, mapping, normalized),
                        }
                    )

                elif logical_name == "trozado":
                    log_code = normalize_log_code(mapping.get("codigo_de_la_troza"))
                    parent_code = parent_tree_from_log_code(log_code)
                    if not (log_code and parent_code):
                        raise ValueError("missing_log_code")
                    raw_species = clean_text(mapping.get("especie"))
                    normalized_species = canonicalize_species(raw_species, canonical_species or {})
                    normalized = {
                        "titulo_habilitante_norm": title,
                        "parcela_corta_norm": parcela,
                        "codigo_troza_norm": log_code,
                        "codigo_arbol_padre_norm": parent_code,
                        "composite_tree_key": composite_tree_key(title, parcela, parent_code),
                        "composite_log_key": composite_log_key(title, parcela, log_code),
                        "especie_norm": normalized_species,
                        "fecha_operacion": normalize_date(mapping.get("fecha")),
                        "volumen_text": normalize_decimal(mapping.get("volumen_m3") or mapping.get("volumen")),
                        "r_private_value": clean_text(mapping.get("r")),
                    }
                    result["logs"].append(
                        {
                            "titulo_habilitante_original": title,
                            "parcela_corta_original": parcela,
                            "codigo_troza_original": clean_text(mapping.get("codigo_de_la_troza")),
                            "especie_original": raw_species,
                            "diametro_mayor_text": normalize_decimal(mapping.get("diametro_mayor_m")) if clean_text(mapping.get("diametro_mayor_m")) else None,
                            "diametro_menor_text": normalize_decimal(mapping.get("diametro_menor_m")) if clean_text(mapping.get("diametro_menor_m")) else None,
                            "longitud_text": normalize_decimal(mapping.get("longitud_aprovechable_m")) if clean_text(mapping.get("longitud_aprovechable_m")) else None,
                            "observaciones_private": clean_text(mapping.get("observaciones")),
                            **normalized,
                            **_source_payload(relative_path, sheet_name, row_number, mapping, normalized),
                        }
                    )

                else:
                    log_code = normalize_log_code(mapping.get("codigo_de_troza"))
                    parent_code = parent_tree_from_log_code(log_code)
                    if not (log_code and parent_code):
                        raise ValueError("missing_dispatch_log_code")
                    normalized = {
                        "titulo_habilitante_norm": title,
                        "parcela_corta_norm": parcela,
                        "codigo_troza_norm": log_code,
                        "codigo_arbol_padre_norm": parent_code,
                        "composite_tree_key": composite_tree_key(title, parcela, parent_code),
                        "composite_log_key": composite_log_key(title, parcela, log_code),
                        "codigo_despacho_norm": clean_text(mapping.get("codigo_de_despacho")),
                        "numero_gtf_norm": normalize_gtf(mapping.get("numero_de_gtf")),
                        "fecha_operacion": normalize_date(mapping.get("fechas") or mapping.get("fecha")),
                    }
                    result["dispatches"].append(
                        {
                            "titulo_habilitante_original": title,
                            "parcela_corta_original": parcela,
                            "codigo_troza_original": clean_text(mapping.get("codigo_de_troza")),
                            "codigo_despacho_original": clean_text(mapping.get("codigo_de_despacho")),
                            "numero_gtf_original": clean_text(mapping.get("numero_de_gtf")),
                            "observaciones_private": clean_text(mapping.get("observaciones")),
                            **normalized,
                            **_source_payload(relative_path, sheet_name, row_number, mapping, normalized),
                        }
                    )
            except ValueError as exc:
                result["rejected"].append(
                    {
                        "reason": str(exc),
                        "sheet": sheet_name,
                        "source_row_number": row_number,
                        "row": mapping,
                    }
                )

    return result


def extract_balance_lines(text: str) -> list[str]:
    lines = [clean_text(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    started = False
    buffer = ""
    parsed: list[str] = []

    for line in lines:
        if "Producto U. Medida Especie Autorizado Extraido Saldo" in line:
            started = True
            continue
        if not started:
            continue
        if line.startswith("Fecha de consulta"):
            continue
        if line.startswith("REGISTRO DE RESOLUCIONES"):
            continue
        buffer = f"{buffer} {line}".strip()
        if re.search(r"-?\d+\.\d{3}\s+-?\d+\.\d{3}\s+-?\d+\.\d{3}$", buffer):
            parsed.append(buffer)
            buffer = ""

    return parsed


def parse_balance_pdf(
    path: Path,
    canonical_species: dict[str, str] | None = None,
    repo_root: Path | None = None,
) -> dict[str, Any]:
    repo_root = repo_root or resolve_repo_root()
    relative_path = path.relative_to(repo_root).as_posix()
    reader = PdfReader(str(path))
    loaded: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    extracted_text: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        extracted_text.append({"page": page_number, "text": text})

        title_match = re.search(
            r"(GOREMAD-GRRNYGA-DRFFS/DFFS-TAH/P-MAD-D-\s*0*08-15)",
            text,
        )
        parcel_match = re.search(r"\(PC\s*([0-9]+)\)", text)

        title = normalize_title(title_match.group(1).replace(" ", "")) if title_match else None
        parcela = normalize_parcela(f"PC {parcel_match.group(1)}") if parcel_match else None

        if not title or not parcela:
            raise ValueError(f"{relative_path}: no se pudo interpretar titulo/parcela del PDF")

        for line in extract_balance_lines(text):
            match = re.match(
                r"(?P<product>.+?)\s+Metros Cúbicos\s+(?P<species>.+?)\s+"
                r"(?P<authorized>-?\d+\.\d{3})\s+"
                r"(?P<extracted>-?\d+\.\d{3})\s+"
                r"(?P<balance>-?\d+\.\d{3})$",
                line,
            )
            if not match:
                rejected.append(
                    {
                        "reason": "line_not_unequivocal",
                        "page": page_number,
                        "source_fragment": line,
                    }
                )
                continue

            raw_species = normalize_species(match.group("species"))
            canonical_species_value = canonicalize_species(
                raw_species,
                canonical_species or {},
            )
            normalized = {
                "titulo_habilitante_norm": title,
                "parcela_corta_norm": parcela,
                "product_type": clean_text(match.group("product")),
                "especie_norm": canonical_species_value,
                "volumen_autorizado_text": normalize_decimal(match.group("authorized")),
                "volumen_extraido_text": normalize_decimal(match.group("extracted")),
                "saldo_reportado_text": normalize_decimal(match.group("balance")),
                "composite_balance_key": composite_balance_key(
                    title,
                    parcela,
                    canonical_species_value,
                ),
            }
            loaded.append(
                {
                    "titulo_habilitante_original": title,
                    "parcela_corta_original": parcela,
                    "especie_original": raw_species,
                    "source_fragment": line,
                    "source_page_number": page_number,
                    **normalized,
                    **_source_payload(
                        relative_path,
                        f"page:{page_number}",
                        page_number,
                        {"line": line},
                        normalized,
                    ),
                }
            )

    return {"loaded": loaded, "rejected": rejected, "extracted_text": extracted_text}
